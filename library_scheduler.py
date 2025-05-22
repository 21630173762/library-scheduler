import json
import os
import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def load_shifts(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_students(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_all_schedules(class_dir):
    schedules = {}
    if not os.path.exists(class_dir):
        print(f"课表文件夹 {class_dir} 不存在！")
        return schedules
    for fname in os.listdir(class_dir):
        if fname.endswith('.json'):
            with open(os.path.join(class_dir, fname), 'r', encoding='utf-8') as f:
                data = json.load(f)
                schedules[data['student_id']] = data['busy']
    return schedules

def is_free(busy_list, day, start, end):
    for b in busy_list:
        if b['day'] == day:
            if not (end <= b['start'] or start >= b['end']):
                return False
    return True

def parse_shift(shift_str):
    import re
    m = re.match(r'(周[一二三四五六日])(?:（|\()(\d{1,2}:\d{2}-\d{1,2}:\d{2})(?:）|\))', shift_str)
    if m:
        return m.group(1), m.group(2)
    return None, None

def get_month_days(year, month):
    days = []
    cal = calendar.Calendar()
    for day in cal.itermonthdates(year, month):
        if day.month == month:
            days.append({
                'date': day,
                'weekday': '周' + '一二三四五六日'[day.weekday()]
            })
    return days

def arrange_month(shifts, students, schedules, year, month):
    days = get_month_days(year, month)
    student_count = {stu['student_id']: 0 for stu in students}
    result = []
    
    # 为前台班创建固定班次与学生的映射
    front_desk_assignments = {}
    library_assignments = {}
    
    # 按照日期和班次整理
    for d in days:
        for shift in shifts:
            shift_day, shift_time = parse_shift(shift['班次'])
            if shift_day != d['weekday']:
                continue
            
            start, end = shift_time.split('-')
            assigned = []
            shift_key = shift['班次']
            
            # 确定使用哪个映射表
            assignment_map = front_desk_assignments if shift.get('类型', '') == '前台班' else library_assignments
            
            # 如果此班次已有固定学生，优先选择他们
            preferred_students = assignment_map.get(shift_key, [])
            candidates = []
            
            # 先尝试使用已分配的固定学生
            for student_id in preferred_students:
                for stu in students:
                    if stu['student_id'] == student_id:
                        busy = schedules.get(stu['student_id'], [])
                        if is_free(busy, shift_day, start, end):
                            assigned.append(stu['name'])
                            student_count[stu['student_id']] += 1
                            if len(assigned) == shift['需求人数']:
                                break
            
            # 如果固定学生不够或不可用，寻找新学生
            if len(assigned) < shift['需求人数']:
                # 按照工作量排序学生
                for stu in sorted([s for s in students if s['student_id'] not in [id for id in preferred_students if id in [s['student_id'] for s in students]]], 
                                 key=lambda x: student_count[x['student_id']]):
                    if len(assigned) == shift['需求人数']:
                        break
                    
                    busy = schedules.get(stu['student_id'], [])
                    if is_free(busy, shift_day, start, end):
                        assigned.append(stu['name'])
                        student_count[stu['student_id']] += 1
                        
                        # 记录这个学生为这个班次的固定学生（如果没有已分配）
                        if shift_key in assignment_map and stu['student_id'] not in assignment_map[shift_key]:
                            assignment_map[shift_key].append(stu['student_id'])
                        elif shift_key not in assignment_map:
                            assignment_map[shift_key] = [stu['student_id']]
            
            result.append({
                '日期': d['date'].strftime('%Y-%m-%d'),
                'weekday': d['weekday'],
                '班次': shift['班次'],
                '学生': ','.join(assigned) if assigned else '/',
                '类型': shift.get('类型', '未知')  # 添加班次类型
            })
    
    return result

def group_days_by_week(days):
    """将日期按周分组，每组是一个周（周一到周日）"""
    weeks = []
    current_week = []
    
    # 确保按日期排序
    days = sorted(days, key=lambda x: x['date'])
    
    for day in days:
        date = day['date']
        # 如果是周一且当前周有内容，开始新一周
        if date.weekday() == 0 and current_week:
            weeks.append(current_week)
            current_week = []
        current_week.append(day)
    
    if current_week:
        weeks.append(current_week)
    
    return weeks

def format_week_title(week):
    """格式化周标题，如 5.1—5.7"""
    if not week:
        return ""
    start_date = week[0]['date']
    end_date = week[-1]['date']
    return f"{start_date.month}.{start_date.day}—{end_date.month}.{end_date.day}"

def get_student_for_week_shift(data, shift, week):
    """获取某一周某个班次的学生名单"""
    # 收集该周所有日期
    week_dates = [day['date'] for day in week]
    
    # 提取班次的日期部分(周X)
    shift_day, _ = parse_shift(shift)
    
    # 检查该周中是否有该班次对应的天
    students_by_day = {}
    for day in week:
        if day['weekday'] == shift_day:
            date_str = day['date'].strftime('%Y-%m-%d')
            # 寻找对应的排班数据
            for item in data:
                if item['班次'] == shift and item['日期'] == date_str:
                    students_by_day[date_str] = item['学生']
    
    # 如果该周没有找到对应的排班，返回"/"
    if not students_by_day:
        return "/"
    
    # 如果找到了多条记录，取第一条非"/"的记录
    for _, student in sorted(students_by_day.items()):
        if student != "/":
            return student
    
    # 默认返回"/"
    return "/"

def sort_shifts(shifts_list):
    """将班次按照周几和时间排序"""
    # 中文周几的顺序映射
    weekday_order = {'周一': 1, '周二': 2, '周三': 3, '周四': 4, '周五': 5, '周六': 6, '周日': 7}
    
    # 解析班次，提取周几和时间
    def extract_shift_info(shift):
        weekday, time_range = parse_shift(shift)
        # 如果解析失败，返回默认值
        if not weekday or not time_range:
            return ('', '')
        return (weekday, time_range)
    
    # 自定义排序
    def shift_sort_key(shift):
        weekday, time_range = extract_shift_info(shift)
        # 首先按周几排序，然后按时间段排序
        return (weekday_order.get(weekday, 999), time_range)
    
    # 返回排序后的班次
    return sorted(shifts_list, key=shift_sort_key)

def create_schedule_sheet(wb, sheet_name, title, data, all_shifts, weeks, year, month):
    """创建一个排班表sheet，在不同周几的班次之间添加空白行增加可读性"""
    ws = wb.create_sheet(title=sheet_name)
    
    # 按周和时间段排序班次
    all_shifts = sort_shifts(all_shifts)
    
    # 设置标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(weeks)+1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(bold=True, size=14)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    for i in range(len(weeks)):
        col_letter = chr(66 + i)  # B, C, D...
        ws.column_dimensions[col_letter].width = 15
    
    # 周标题行
    for i, week in enumerate(weeks):
        if not week:
            continue
        week_title = format_week_title(week)
        cell = ws.cell(row=2, column=i+2, value=week_title)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    
    # 提取每个班次的周几
    weekday_shifts = {}  # {周一: [班次1, 班次2...], 周二: [班次1, 班次2...], ...}
    for shift in all_shifts:
        weekday, _ = parse_shift(shift)
        if weekday not in weekday_shifts:
            weekday_shifts[weekday] = []
        weekday_shifts[weekday].append(shift)
    
    # 按顺序填写班次及空白行
    row_index = 3  # 从第3行开始
    weekday_order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    
    for weekday in weekday_order:
        if weekday not in weekday_shifts:
            continue
        
        # 当前星期的班次
        curr_shifts = weekday_shifts[weekday]
        
        # 填写当前星期的班次
        for shift in curr_shifts:
            # 填写班次名称
            ws.cell(row=row_index, column=1, value=shift)
            
            # 填写排班数据
            for j, week in enumerate(weeks):
                col = j + 2
                student = get_student_for_week_shift(data, shift, week)
                cell = ws.cell(row=row_index, column=col, value=student)
                cell.alignment = Alignment(horizontal='center')
            
            row_index += 1
        
        # 在不同周几之间添加空白行（最后一个周几后不加）
        if weekday != weekday_order[-1] and any(wd in weekday_shifts for wd in weekday_order[weekday_order.index(weekday)+1:]):
            # 添加空白行
            row_index += 1
    
    # 添加表格边框
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # 计算最大行数
    max_row = row_index - 1
    max_col = len(weeks) + 1
    
    for row in range(1, max_row + 1):
        # 跳过空白行
        if not ws.cell(row=row, column=1).value and row > 2:
            continue
            
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    return ws

def save_to_xlsx_with_weeks(data, filename, year, month):
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认的Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 获取该月所有日期
    all_days = get_month_days(year, month)
    
    # 按周分组日期
    weeks = group_days_by_week(all_days)
    
    # 分离前台班和书库班数据
    front_desk_data = []
    library_data = []
    
    # 根据类型分类数据
    for item in data:
        if "前台" in item.get('类型', ''):
            front_desk_data.append(item)
        else:
            library_data.append(item)
    
    # 获取所有班次（不需要排序，在create_schedule_sheet中会排序）
    front_desk_shifts_list = list(set([row['班次'] for row in front_desk_data]))
    library_shifts_list = list(set([row['班次'] for row in library_data]))
    
    # 创建前台班sheet
    create_schedule_sheet(wb, "前台班排班表", f"前台班排班表（{month}月）", 
                          front_desk_data, front_desk_shifts_list, weeks, year, month)
    
    # 创建书库班sheet
    create_schedule_sheet(wb, "书库班排班表", f"书库班排班表（{month}月）", 
                          library_data, library_shifts_list, weeks, year, month)
    
    # 保存
    wb.save(filename)

if __name__ == '__main__':
    year = 2025
    month = 5
    student_file = 'student.txt'
    class_dir = 'class'
    shift_files = ['前台班班次.txt', '书库班班次.txt']
    students = load_students(student_file)
    schedules = load_all_schedules(class_dir)
    
    # 加载不同类型的班次
    front_desk_shifts = []
    library_shifts = []
    
    for f in shift_files:
        shifts = load_shifts(f)
        # 根据文件名标记班次类型
        for shift in shifts:
            if "前台" in f:
                shift['类型'] = '前台班'
                front_desk_shifts.append(shift)
            else:
                shift['类型'] = '书库班'
                library_shifts.append(shift)
    
    # 合并所有班次
    shifts = front_desk_shifts + library_shifts
    
    result = arrange_month(shifts, students, schedules, year, month)
    save_to_xlsx_with_weeks(result, f'{month}月排班表.xlsx', year, month)
    print(f"{month}月排班表已生成！") 